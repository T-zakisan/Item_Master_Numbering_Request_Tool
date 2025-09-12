# test_04(更新_9).py
import json, os, re, csv
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkfont
from datetime import datetime

YOUR_APP_NAME="DennoFactoryApp"
SCHEMA_PATH = r"V:\00.各個人用フォルダ\999.yamazaki\test\field_schema.csv"   # スキーマCSVの既定パス（共有フォルダ）
EMAIL_DOMAIN="@hantak.co.jp"
DEPTS=["資材",  "部品センター",  "技術",  "技術受付",  "採番担当"]
MENU_ITEMS=["新規",  "保存",  "取込",  "採番依頼",  "採番完了",  "CSV転送",  "設定"]
ACL={
    "資材":         {"新規":1,  "保存":1,  "取込":1,  "採番依頼":1,  "採番完了":0,  "CSV転送":0,  "設定":1},
    "部品センター": {"新規":1,  "保存":1,  "取込":1,  "採番依頼":1,  "採番完了":0,  "CSV転送":0,  "設定":1},
    "技術":         {"新規":1,  "保存":1,  "取込":1,  "採番依頼":1,  "採番完了":0,  "CSV転送":0,  "設定":1},
    "技術受付":     {"新規":0,  "保存":0,  "取込":1,  "採番依頼":0,  "採番完了":0,  "CSV転送":1,  "設定":1},
    "採番担当":     {"新規":1,  "保存":0,  "取込":1,  "採番依頼":0,  "採番完了":1,  "CSV転送":0,  "設定":1},
}
FONT_PRESETS={10:"小",14:"中",18:"大"}
ROW_PAD_Y = 8  # 右欄の行間（pady）をまとめて制御
DIGITS_RE = re.compile(r"^\d+$")
def settings_path()->Path:
    base=Path(os.environ.get("APPDATA","."))
    base/=YOUR_APP_NAME; base.mkdir(parents=True,exist_ok=True)
    return base/"settings.json"

def load_settings():
    p=settings_path()
    if not p.exists(): return None
    try:
        with p.open("r",encoding="utf-8") as f: data=json.load(f)
        if data.get("department") not in DEPTS: return None
        if not data.get("email","").endswith(EMAIL_DOMAIN): return None
        if not data.get("output_dir"): return None
        fs=int(data.get("font_size",12))
        data["font_size"]=fs if fs in (10,14,18) else 12
        data["maximize_on_start"]=bool(data.get("maximize_on_start", False))
        return data
    except: return None

def save_settings(d:dict):
    with settings_path().open("w",encoding="utf-8") as f:
        json.dump(d,f,ensure_ascii=False,indent=2)

def pick_jp_ui_font():
    prefs = ["BIZ UDGothic","BIZ UDゴシック","Yu Gothic UI","Yu Gothic","Meiryo UI","Meiryo","MS UI Gothic"]
    try:
        fams=set(tkfont.families())
        for n in prefs:
            if n in fams: return n
    except: pass
    return "TkDefaultFont"


# 既定プルダウン（設定未選択時のフォールバック）
DROPDOWN_DEFAULT={"標準単位":["未設定：［設定］から単位一覧.xlsx を指定すること！"],"売上一覧表用分類":["未設定：［設定］から売上一覧用分類.xlsx を指定すること！"]}

ASCII_RE=re.compile(r"[ -~\uFF61-\uFF9F]*")  # 半角英数記号 + 半角カタカナ
def is_ascii_symbol(s:str)->bool:
    return ASCII_RE.fullmatch(s or "") is not None
FLOAT_RE=re.compile(r"^\d+(\.\d+)?$")
def is_float_str(s:str)->bool:
    return FLOAT_RE.fullmatch(s or "") is not None
EMAIL_LOCAL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+$")
def is_email_local(s:str)->bool:
    return EMAIL_LOCAL_RE.fullmatch(s or "") is not None


SEIZO_LABELS = {
    "0": "0：ﾛｯﾄ手配品目",
    "3": "3：ﾏﾆｭｱﾙ手配品目",
}

def trunc31(s:str)->str:
    s=s or ""
    return s if len(s)<=31 else (s[:30]+"⋯")

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(YOUR_APP_NAME)
        self.geometry("980x560"); self.minsize(820,480)

        fam=pick_jp_ui_font()
        self.app_font=tkfont.Font(family=fam,size=12)
        self.bold_font=tkfont.Font(family=fam,size=12,weight="bold")
        self.btn_font=tkfont.Font(family=fam,size=12)

        # 2系統フォント
        self.font_mono = tkfont.Font(family="Consolas", size=12)
        self.font_jp   = tkfont.Font(family=pick_jp_ui_font(), size=12)
        self.font_jp_b = tkfont.Font(family=self.font_jp.cget("family"), size=12, weight="bold")

        # スタイル
        self.style = ttk.Style()
        self.style.configure("Mono.TLabel", font=self.font_mono)
        self.style.configure("Mono.TEntry",  font=self.font_mono)
        self.style.configure("Mono.TCombobox", font=self.font_mono)  # 使うなら
        self.style.configure("Jp.TLabel",   font=self.font_jp)
        self.style.configure("JpBold.TLabel", font=self.font_jp_b)
        self.style.configure("Jp.TButton",  font=self.font_jp, padding=(10,5))
        self.style.configure("Jp.TRadiobutton", font=self.font_jp)


        self.dim_font = tkfont.Font(
            family=self.font_jp.cget("family"),
            size=max(10, self.font_jp['size'] - 1)
        )
        self.style.configure("Dim.TLabel", font=self.dim_font, foreground="#666666")


        # 既定は等幅で統一したい場（データ側など）
        self.option_add("*Font", self.font_jp)      # 全体の日本語UI
        self.option_add("*Entry.Font", self.font_mono)
        self.option_add("*Listbox.Font", self.font_mono)

        # toolbar
        self.toolbar=ttk.Frame(self,padding=4); self.toolbar.pack(side="top",fill="x")
        self.buttons={}
        for item in MENU_ITEMS:
            cmd = {
                "設定": self.open_settings,
                "新規": self.show_new_view,
                "保存": self.on_save_csv,
                "取込": self.on_import_csv,
                "採番依頼": self.on_request_csv,  # ← 追加
            }.get(item, self.dummy)
            b=ttk.Button(self.toolbar,text=item,command=cmd,style="Jp.TButton"); b.pack(side="left",padx=2)
            self.buttons[item]=b

        self.work=ttk.Frame(self); self.work.pack(fill="both",expand=True)

        self.rows=[]; self.cur_index=None
        self.FIELDS = []  # [(new_label, old_label, csvkey)]
        # プルダウン値（設定で上書き）
        self.dropdown_values = DROPDOWN_DEFAULT.copy()   # { "標準単位":[..], "売上一覧表用分類":[..] }
        # 表示→CSV値の変換マップ（将来のCSV出力用に保持。今は未使用）
        self.dropdown_export_map = {"標準単位":{}, "売上一覧表用分類":{}}
        self.settings=load_settings()
        if self.settings is None:
            self.after(100,self.open_settings)
        else:
            self.apply_settings()
            self.load_field_defs()
            self.load_dropdown_masters()
            self.after(50,self.auto_resize_after_settings)
            self.show_new_view()

        # ボタンのショートカット
        # self.bind_all("<Return>",           self._on_return,    add="+")
        self.bind_all("<Control-Return>",       self._sc_add)
        self.bind_all("<Delete>",               self._sc_del)
        self.bind_all("<Control-Key-space>",    self._sc_upd)
        self.bind_all("<Escape>",               self._sc_clr)
        self.bind_all("<Control-s>",            self._sc_save)
        self.bind_all("<Control-o>",            self._sc_import)

    def shortcuts_enabled(self)->bool:
        return bool((self.settings or {}).get("enable_shortcuts", False))

    def _sc_add(self, e=None):
        if not self.shortcuts_enabled(): return
        self.on_add()

    def _sc_del(self, e=None):
        if not self.shortcuts_enabled(): return
        self.on_delete()

    def _sc_upd(self, e=None):
        if not self.shortcuts_enabled(): return
        self.on_update()

    def _sc_clr(self, e=None):
        if not self.shortcuts_enabled(): return
        self.on_clear()

    def _sc_save(self, e=None):
        if not self.shortcuts_enabled(): return
        self.on_save_csv()

    def _sc_import(self, e=None):
        if not self.shortcuts_enabled(): return
        self.on_import_csv()




    # # 項目定義CSVの読み込み（列A=新, 列B=旧, 列C=CSVキー）。ヘッダなし前提。
    def load_field_defs(self):
        if not self.load_schema():
            messagebox.showerror("エラー", f"スキーマCSVが見つかりません:\n{SCHEMA_PATH}")
            self.after(100, self.destroy); return
        self.FIELDS   = [(r["項目名"], r.get("旧表記",""), r["項目名"])
                        for r in self.schema_rows if r.get("項目名")]
        self.OUT_COLS = [r["項目名"] for r in self.schema_rows if r.get("項目名")]
        if not self.FIELDS:
            messagebox.showerror("エラー", "スキーマに『項目名』列が読めません。区切りや余計なタブ・空白を確認してください。")
            self.after(100, self.destroy)

    @staticmethod
    def _open_schema(path: Path):
        import csv
        enc = None; head = ""
        for e in ("utf-8-sig","cp932"):
            try:
                with path.open("r", encoding=e, newline="") as f:
                    head = f.readline()
                enc = e; break
            except UnicodeError:
                continue
        if enc is None:
            raise RuntimeError("schema encoding not readable")

        delim = "\t" if ("\t" in head and "," not in head) else ","

        rows = []
        with path.open("r", encoding=enc, newline="") as f:
            rdr = csv.reader(f, delimiter=delim)
            try:
                headers = next(rdr)
            except StopIteration:
                return []
            headers = [(h or "").strip() for h in headers]

            for rec in rdr:
                d = {}
                for k, v in zip(headers, rec):
                    kk = (k or "").strip()
                    if not kk or kk.startswith("Unnamed"):
                        continue
                    d[kk] = (v or "").strip()
                rows.append(d)
        return rows



    def get_schema_path(self):
        return Path(SCHEMA_PATH)

    def load_schema(self)->bool:
        p = self.get_schema_path()
        self.schema_rows = []
        if not p or not p.exists():
            messagebox.showwarning("注意", f"スキーマCSVが見つかりません:\n{p}\nUI定義はフォールバックします。")
            return False
        try:
            self.schema_rows = self._open_schema(p)
            return bool(self.schema_rows)
        except Exception as e:
            messagebox.showerror("エラー", f"スキーマCSVの読込に失敗:\n{p}\n{e}")
            return False


    def _dept_col(self, prefix:str, dept:str)->str:
        return f"{dept}_{prefix}"

    def schema_get_mode(self, dept:str, key:str)->str:
        col = self._dept_col("mode", dept)
        for r in self.schema_rows:
            if r.get("項目名")==key:
                v = (r.get(col) or "").lower()
                return v or "entry"
        return "entry"  # スキーマ未定義時の最小フォールバック

    def schema_is_required(self, dept:str, key:str)->bool:
        col = self._dept_col("required", dept)
        for r in self.schema_rows:
            if r.get("項目名")==key:
                v = (r.get(col) or "").lower()
                return v in ("1","true")
        return False

    def schema_default(self, dept:str, key:str)->str:
        col = self._dept_col("default", dept)
        for r in self.schema_rows:
            if r.get("項目名")==key:
                return r.get(col,"")
        return ""




    # ［新規］の左欄の縞々
    def _zebra_listbox(self):
        if not hasattr(self, "listbox"): return
        lb = self.listbox
        for i in range(lb.size()):
            bg = "#FFFFFF" if (i % 2) == 0 else "#F3F3F3"  # 偶数:白 / 奇数:薄灰
            try:
                lb.itemconfigure(i, background=bg)
            except tk.TclError:
                pass  # 古いTkで未対応なら無視

    # Excelマスタの読込（openpyxl があれば使用）
    def load_dropdown_masters(self):
        self.dropdown_values = DROPDOWN_DEFAULT.copy()
        self.dropdown_export_map = {"標準単位":{}, "売上一覧表用分類":{}}
        try:
            import openpyxl  # 依存が無い場合は except へ
        except Exception:
            return
        s = self.settings or {}
        # 標準単位
        upath = s.get("units_master","")
        if upath and Path(upath).exists():
            try:
                wb = openpyxl.load_workbook(upath, data_only=True)
                ws = wb.active
                vals=[]; mapping={}
                for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if i==1:  # ヘッダ行想定
                        continue
                    b = (row[1] or "") if len(row)>1 else ""
                    if b:
                        # 「プルダウン表示, csv」形式に対応
                        if "," in str(b):
                            disp, csvv = [x.strip() for x in str(b).split(",",1)]
                        else:
                            disp, csvv = str(b).strip(), str(b).strip()
                        vals.append(disp); mapping[disp]=csvv
                if vals:
                    self.dropdown_values["標準単位"]=vals
                    self.dropdown_export_map["標準単位"]=mapping
            except Exception:
                pass
        # 売上一覧表用分類
        spath = s.get("salescat_master","")
        if spath and Path(spath).exists():
            try:
                wb = openpyxl.load_workbook(spath, data_only=True)
                ws = wb.active
                vals=[]; mapping={}
                for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if i==1:
                        continue
                    a = (row[0] or "") if len(row)>0 else ""
                    b = (row[1] or "") if len(row)>1 else ""
                    disp = str(b).strip()
                    csvv = str(a).strip()
                    if disp:
                        vals.append(disp)
                        if csvv:
                            mapping[disp]=csvv
                if vals:
                    self.dropdown_values["売上一覧表用分類"]=vals
                    self.dropdown_export_map["売上一覧表用分類"]=mapping
            except Exception:
                pass


    def _collect_form_row(self)->dict:
        dept = (self.settings or {}).get("department","資材")
        row = {}
        for _,_,csvkey in self.FIELDS:
            if csvkey=="製造手配区分" and (csvkey+"_raw") in self.vars:
                row[csvkey] = self.vars[csvkey+"_raw"]
            else:
                row[csvkey] = self.vars[csvkey].get() if csvkey in self.vars else ""
            if not row[csvkey]:
                dv = self.schema_default(dept, csvkey)
                if dv: row[csvkey] = dv
        return row

    def _disp_text(self, row:dict)->str:
        disp_full = row.get("品目名1") or row.get("品目名2") or row.get("品目ｺｰﾄﾞ") or "(無題)"
        return trunc31(disp_full)

    def dummy(self): pass

    def apply_fontsize(self, pt:int):
        for f in (self.font_mono, self.font_jp, self.font_jp_b):
            f.configure(size=pt)
        self.refresh_fonts()

    def refresh_fonts(self):
        if hasattr(self,"listbox"):
            try: self.listbox.config(font=self.font_mono)
            except: pass

    def apply_settings(self):
        s=self.settings
        self.apply_fontsize(int(s.get("font_size",12)))
        dept=s.get("department")
        for item,b in self.buttons.items():
            b.config(state=("normal" if ACL.get(dept,{}).get(item,0) else "disabled"))

    def auto_resize_after_settings(self):
        if self.settings.get("maximize_on_start", False):
            try: self.state("zoomed"); return
            except: pass
        self.update_idletasks()
        total=8
        for b in self.buttons.values(): total+=b.winfo_reqwidth()+4
        height=max(480,self.toolbar.winfo_reqheight()+420); width=max(820,total)
        self.geometry(f"{width}x{height}")

    # 設定
    def open_settings(self):
        dlg=tk.Toplevel(self); dlg.title("設定"); dlg.transient(self); dlg.grab_set()
        frm=ttk.Frame(dlg,padding=12); frm.pack(fill="both",expand=True)
        frm.option_add("*Font", self.font_jp)
        r = 0

        # 色（不正時の黄色化に使用）
        _tmp=tk.Entry(frm); _default_bg=_tmp.cget("bg"); _tmp.destroy()
        _invalid_bg="#FFF29A"
        def _bind_validate(widget, rule, getter):
            def _on_blur(_evt=None):
                v=getter()
                ok=True
                if rule=="email":
                    ok=is_email_local(v)
                elif rule=="digits_required":
                    ok=bool(v) and DIGITS_RE.fullmatch(v or "")
                elif rule=="dir":
                    ok=Path(v).exists()
                widget.config(bg=(_default_bg if ok else _invalid_bg))
            widget.bind("<FocusOut>", _on_blur)
            return _on_blur


        # メール（1行：ラベル｜入力欄｜@ドメイン）
        ttk.Label(frm,text="■メールアドレス",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        email_var=tk.StringVar()
        email_ent=tk.Entry(frm,textvariable=email_var,width=15,font=self.font_mono)
        email_ent.grid(row=r,column=0,sticky="w",pady=(0,3*ROW_PAD_Y))
        ttk.Label(frm,text=EMAIL_DOMAIN,font=self.font_mono).grid(row=r,column=1,sticky="w",padx=(0,0),pady=(0,3*ROW_PAD_Y))
        _val_email=_bind_validate(email_ent,"email",lambda: email_var.get())


        r += 1
        # 社員番号（1行：ラベル｜入力欄）
        ttk.Label(frm,text="■社員番号",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        empno_var=tk.StringVar()
        r += 1
        empno_ent = tk.Entry(frm, textvariable=empno_var, width=15, font=self.font_mono)
        empno_ent.grid(row=r,column=0,sticky="w",pady=(0,3*ROW_PAD_Y))
        _val_emp=_bind_validate(empno_ent,"digits_required",lambda: empno_var.get())
        r += 1


        # 所属
        ttk.Label(frm,text="■所属",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r +=1
        dept_var=tk.StringVar(value=(self.settings or {}).get("department",DEPTS[0]))
        box=ttk.Frame(frm); box.grid(row=r,column=0,columnspan=2,sticky="w",pady=(0,3*ROW_PAD_Y))
        for d in DEPTS:
            ttk.Radiobutton(box,text=d,value=d,variable=dept_var,style="Jp.TRadiobutton").pack(anchor="w")
        r += 1

        # 表示サイズ
        ttk.Label(frm,text="■表示サイズ",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        fs_var=tk.IntVar(value=int((self.settings or {}).get("font_size",12)))
        box_fs=ttk.Frame(frm); box_fs.grid(row=r,column=0,columnspan=2,sticky="w",pady=(0,3*ROW_PAD_Y))
        for pt,label in [(10,"小"),(14,"中"),(18,"大")]:
            ttk.Radiobutton(box_fs,text=f"{label}（{pt}pt）",value=pt,variable=fs_var,style="Jp.TRadiobutton").pack(anchor="w")
        r += 1

        # 出力場所
        ttk.Label(frm,text="■出力場所",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        out_var=tk.StringVar(value=(self.settings or {}).get("output_dir",""))
        def choose_dir():
            d=filedialog.askdirectory(title="■出力先フォルダ選択")
            if d:
                out_var.set(d)
                lbl_out.config(text=d)

        ttk.Button(frm,text="フォルダ選択",command=choose_dir,style="Jp.TButton").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))

        lbl_out = ttk.Label(frm, text=out_var.get(),
                    wraplength=500, justify="left",
                    style="Jp.TLabel", font=self.font_mono)
                    # style="Jp.TLabel")
        lbl_out.grid(row=r, column=1, columnspan=2, sticky="w", pady=(0,ROW_PAD_Y))
        r += 1


        # プルダウンマスタ
        ttk.Label(frm,text="■プルダウンマスタ",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        # 標準単位
        ttk.Label(frm,text="標準単位",style="Jp.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        units_var = tk.StringVar(value=(self.settings or {}).get("units_master",""))
        def _fmt_path(p):
            if not p: return ""
            mtime = datetime.fromtimestamp(Path(p).stat().st_mtime).strftime("%Y.%m.%d")
            return f"{Path(p).name} @ {mtime}"
        ttk.Button(frm,text="ファイル選択",
                    command=lambda: (lambda p=filedialog.askopenfilename(
                        title="標準単位マスタを選択", filetypes=[("Excel","単位一覧*.xlsx")]):
                        units_var.set(p) if p else None)(),
                    style="Jp.TButton").grid(row=r,column=1,sticky="w",pady=(0,ROW_PAD_Y))
        units_lbl = ttk.Label(frm,text=_fmt_path(units_var.get()), font=self.font_mono)
        units_lbl.grid(row=r,column=2,sticky="w",pady=(0,ROW_PAD_Y))
        def _upd_units_lbl(*_):
            units_lbl.config(text=_fmt_path(units_var.get()))
        units_var.trace_add("write", _upd_units_lbl)
        r += 1

        # 売上一覧表用分類
        ttk.Label(frm,text="売上一覧表用分類",style="Jp.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        sales_var = tk.StringVar(value=(self.settings or {}).get("salescat_master",""))
        ttk.Button(frm,text="ファイル選択",
                    command=lambda: (lambda p=filedialog.askopenfilename(
                        title="売上一覧表用分類マスタを選択", filetypes=[("Excel","売上一覧用分類*.xlsx")]):
                        sales_var.set(p) if p else None)(),
                    style="Jp.TButton").grid(row=r,column=1,sticky="w",pady=(0,ROW_PAD_Y))
        sales_lbl = ttk.Label(frm,text=_fmt_path(sales_var.get()),font=self.font_mono)
        sales_lbl.grid(row=r,column=2,sticky="w",pady=(0,ROW_PAD_Y))
        def _upd_sales_lbl(*_):
            sales_lbl.config(text=_fmt_path(sales_var.get()))
        sales_var.trace_add("write", _upd_sales_lbl)
        r += 1



        ttk.Label(frm,text="■その他",style="JpBold.TLabel").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        show_old_var=tk.BooleanVar(value=bool((self.settings or {}).get("show_old_alias", False)))
        ttk.Checkbutton(frm,text="旧項目名を併記",variable=show_old_var,style="Jp.TRadiobutton").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        max_var=tk.BooleanVar(value=bool((self.settings or {}).get("maximize_on_start", False)))
        ttk.Checkbutton(frm,text="起動時は最大化",variable=max_var,style="Jp.TRadiobutton").grid(row=r,column=0,sticky="w",pady=(0,ROW_PAD_Y))
        r += 1
        shortcut_var = tk.BooleanVar(value=bool((self.settings or {}).get("enable_shortcuts", False)))
        ttk.Checkbutton(frm, text="ショートカット使用", variable=shortcut_var, style="Jp.TRadiobutton").grid(row=r, column=0, sticky="w", pady=(0, ROW_PAD_Y))
        r += 1

        cur=self.settings or {}
        if cur.get("email","").endswith(EMAIL_DOMAIN):
            email_var.set(cur["email"].replace(EMAIL_DOMAIN,""))
        if cur.get("employee_no"): empno_var.set(cur["employee_no"])


        def _val_dir():
            ok = Path(out_var.get()).exists()
            lbl_out.config(foreground=("#000000" if ok else "#C00000"))
            return ok

        # 初期表示時にも色を反映
        _val_email(); _val_emp(); _val_dir()

        def save_and_close():
            if not re.fullmatch(r"[A-Za-z0-9._%+\-]+",email_var.get()): messagebox.showerror("エラー","メールローカル部は半角英数記号のみ"); return
            if dept_var.get() not in DEPTS: messagebox.showerror("エラー","所属を選択"); return
            if not _val_dir():
                messagebox.showerror("エラー","有効なフォルダを指定"); return
            if not DIGITS_RE.fullmatch(empno_var.get()):
                messagebox.showerror("エラー","社員番号は半角数字のみ、かつ必須です"); return
            # 念のため直前にも色反映
            _val_email(); _val_emp(); _val_dir()


            data={"app_name":YOUR_APP_NAME,
                    "email": email_var.get() + EMAIL_DOMAIN,
                    "department":dept_var.get(),
                    "output_dir":out_var.get(),
                    "employee_no": empno_var.get(),
                    "units_master": units_var.get(),
                    "salescat_master": sales_var.get(),
                    "show_old_alias": bool(show_old_var.get()),
                    "font_size":int(fs_var.get()),
                    "maximize_on_start":bool(max_var.get()),
                    "enable_shortcuts":  bool(shortcut_var.get()),
                }

            save_settings(data)
            self.settings = data
            self.apply_settings()
            self.load_field_defs()
            self.load_dropdown_masters()
            dlg.destroy()
            self.after(50, self.auto_resize_after_settings)
            self.show_new_view()  # ← 設定の反映を即時適用（新旧項目名の併記やフォントなど）
            messagebox.showinfo("情報","設定を保存しました")

        btns=ttk.Frame(frm); btns.grid(row=r,column=0,columnspan=3,sticky="e",pady=(ROW_PAD_Y,0))

        ttk.Button(btns,text="保存",        command=save_and_close, style="Jp.TButton").pack(side="right",padx=4)
        ttk.Button(btns,text="キャンセル",  command=dlg.destroy,    style="Jp.TButton").pack(side="right")

    # 新規ビュー
    def show_new_view(self):
        for w in self.work.winfo_children(): w.destroy()

        body=ttk.Frame(self.work); body.pack(fill="both",expand=True,padx=6,pady=6)

        left=ttk.Frame(body); left.pack(side="left",fill="both",expand=False)
        show_old = bool((self.settings or {}).get("show_old_alias", False))
        # FIELDS から品目名1の新旧を探す
        new_nm, old_nm, _ = next(((n,o,k) for n,o,k in self.FIELDS if k=="品目名1"),("部品名称","", "品目名1"))
        if show_old and old_nm:
            capf = ttk.Frame(left)
            capf.pack(anchor="w")
            ttk.Label(capf,text=new_nm,style="JpBold.TLabel").grid(row=0,column=0,sticky="w")
            ttk.Label(capf,text=old_nm,style="Dim.TLabel").grid(row=0,column=1,sticky="w",padx=(4,0))
        else:
            ttk.Label(left,text=new_nm,style="JpBold.TLabel").pack(anchor="w")

        lf=ttk.Frame(left); lf.pack(fill="both",expand=True)
        self.listbox=tk.Listbox(lf,height=22,width=31, exportselection=False)
        self.listbox.config(font=self.font_mono,
                    selectbackground="#1E88E5",
                    selectforeground="#FFFFFF",
                    activestyle="none")
        self.listbox.pack(side="left",fill="both",expand=True)
        sb=ttk.Scrollbar(lf,orient="vertical",command=self.listbox.yview); sb.pack(side="right",fill="y")
        self.listbox.config(yscrollcommand=sb.set)
        self.listbox.bind("<<ListboxSelect>>",self.on_select_row)


        # 右：フォーム（項目名は太字・右寄せ、入力欄を隣接）
        right=ttk.Frame(body); right.pack(side="left",fill="both",expand=True,padx=(12,0))
        self.vars={}; self.widgets={}; self.validators={}
        dept = (self.settings or {}).get("department","資材")

        # デフォルト背景色を取得（不正時の黄色化と戻し用）
        tmp=tk.Entry(right); default_bg=tmp.cget("bg"); tmp.destroy()
        invalid_bg="#FFF29A"

        def bind_validate(csvkey, widget, rule):
            def _on_blur(_evt=None):
                val=widget.get()
                ok=True
                if rule=="ascii":
                    ok=is_ascii_symbol(val)
                elif rule=="float":
                    ok=is_float_str(val)
                widget.config(bg=(default_bg if ok else invalid_bg))
            widget.bind("<FocusOut>", _on_blur)
            self.validators[csvkey]=_on_blur


        rowi=0
        show_old = bool((self.settings or {}).get("show_old_alias", False))
        for new_label, old_label, csvkey in self.FIELDS:
            mode = self.schema_get_mode(dept, csvkey)
            if mode=="hide": continue

            # ラベル列：新名=太字、旧名=薄字（併記ON時のみ）
            if show_old and old_label:
                labf = ttk.Frame(right)
                labf.grid(row=rowi,column=0,sticky="e",pady=ROW_PAD_Y,padx=(0,8))
                ttk.Label(labf,text=new_label,style="JpBold.TLabel").grid(row=0,column=0,sticky="e")
                ttk.Label(labf,text=old_label,style="Dim.TLabel").grid(row=0,column=1,sticky="w",padx=(4,0))
            else:
                ttk.Label(right, text=new_label, style="JpBold.TLabel").grid(row=rowi,column=0,sticky="e",pady=ROW_PAD_Y,padx=(0,8))

            var=tk.StringVar(); self.vars[csvkey]=var
            default = self.schema_default(dept, csvkey)

            if mode=="fixed":
                # 製造手配区分は表示ラベル化。内部値は *_raw に保持
                if csvkey=="製造手配区分":
                    raw = default
                    self.vars[csvkey+"_raw"] = raw
                    var.set(SEIZO_LABELS.get(raw, raw))
                else:
                    var.set(default)
                w = tk.Label(right, textvariable=var, font=self.font_mono, anchor="w")
                w.grid(row=rowi, column=1, sticky="we", pady=ROW_PAD_Y)
            elif mode=="label":
                w = tk.Label(right, textvariable=var, font=self.font_mono, anchor="w")
                w.grid(row=rowi, column=1, sticky="we", pady=ROW_PAD_Y)
            elif mode=="dropdown":
                values = self.dropdown_values.get(csvkey) or self.dropdown_values.get(new_label) or ["—"]
                # 既定値がCSV値の場合は表示名へ逆引き
                default_disp = default
                if default and values:
                    m = self.dropdown_export_map.get(csvkey) or {}
                    # m: 表示名 → CSV値 なので逆引き
                    rev = {v:k for k,v in m.items()}
                    default_disp = rev.get(default, default)
                    if default_disp and default_disp not in values:
                        values = [default_disp] + values
                w = ttk.Combobox(right, textvariable=var, values=values,
                                state="readonly", style="Mono.TCombobox")
                if default_disp:
                    w.set(default_disp)
                w.grid(row=rowi, column=1, sticky="we", pady=ROW_PAD_Y)

            else:
                # 背景色を変更できるよう tk.Entry を使用
                w=tk.Entry(right,textvariable=var)
                if default: var.set(default)
                w.grid(row=rowi,column=1,sticky="we",pady=ROW_PAD_Y)
                # 入力規則：品目名1/2は半角英数記号のみ。フォーカスアウトで判定し色変更
                if csvkey in ("品目名1","品目名2"):
                    bind_validate(csvkey,w,"ascii")
                elif csvkey=="通常ﾚｽ率":
                    bind_validate(csvkey,w,"float")

            self.widgets[csvkey]=w
            rowi+=1

        # 新規のときは品目コード入力を無効化
        mode_code = self.schema_get_mode(dept, "品目ｺｰﾄﾞ")
        w = self.widgets.get("品目ｺｰﾄﾞ")
        if isinstance(w, tk.Entry):
            w.config(state=("normal" if mode_code == "entry" else "disabled"))

        # 下部操作ボタン
        btns=ttk.Frame(right)
        btns.grid(row=rowi, column=1, sticky="w", pady=(12,0))  # 入力欄と同じ列=1、左寄せ
        ttk.Button(btns,text="追加",    command=self.on_add,    style="Jp.TButton").pack(side="left", padx=4)
        ttk.Button(btns,text="削除",    command=self.on_delete, style="Jp.TButton").pack(side="left", padx=4)
        ttk.Button(btns,text="変更",    command=self.on_update, style="Jp.TButton").pack(side="left", padx=4)
        ttk.Button(btns,text="クリア",  command=self.on_clear,  style="Jp.TButton").pack(side="left", padx=4)

        right.grid_columnconfigure(0,weight=0); right.grid_columnconfigure(1,weight=1)
        self.refresh_fonts()
        self._zebra_listbox()

    # 左選択
    def on_select_row(self, _evt=None):
        if not self.listbox.curselection(): return
        idx=self.listbox.curselection()[0]; self.cur_index=idx; row=self.rows[idx]
        for _,_,csvkey in self.FIELDS:
            if csvkey in self.vars:
                val = row.get(csvkey,"")
                if csvkey=="製造手配区分":
                    # 表示はラベル化、内部 raw を更新
                    self.vars[csvkey+"_raw"] = val
                    self.vars[csvkey].set(SEIZO_LABELS.get(val,val))
                else:
                    self.vars[csvkey].set(val)


    def _get_selected_index(self):
        sel = self.listbox.curselection()
        if sel: return sel[0]
        if self.cur_index is not None and 0 <= self.cur_index < len(self.rows):
            return self.cur_index
        return None



    # カーソルを品目名1へ
    def _focus_name1(self):
        try:
            w = self.widgets.get("品目名1")
            if w:
                w.focus_set()
        except Exception:
            pass

    # [追加]ボタン
    def _on_return(self, e):
        w = self.focus_get()
        try:
            cls = w.winfo_class()
        except Exception:
            cls = ""
        # プルダウンにフォーカスがある間は［追加］を起動しない
        if cls in ("TCombobox", "ttk::combobox"):
            return  # 既定の確定動作だけ通す
        # 必要なら Entry でも無効化したい場合は次行を有効化:
        self._sc_add()

    def on_add(self):
        dept = (self.settings or {}).get("department","資材")
        for _,_,csvkey in self.FIELDS:
            if self.schema_is_required(dept, csvkey) and not self.vars.get(csvkey,"").get().strip():
                messagebox.showerror("エラー", f"必須: {csvkey}"); return
        nm1=self.vars.get("品目名1",tk.StringVar(value="")).get()
        nm2=self.vars.get("品目名2",tk.StringVar(value="")).get()
        if not is_ascii_symbol(nm1) or not is_ascii_symbol(nm2):
            messagebox.showerror("エラー","品目名1/2は半角英数記号のみ"); return

        row = self._collect_form_row()

        # 既存行と完全一致ならスキップ
        cols = [k for _,_,k in self.FIELDS]
        if any(all(r.get(c,"")==row.get(c,"") for c in cols) for r in self.rows):
            messagebox.showinfo("情報","全項目が一致する行があるため追加をスキップしました"); return

        self.rows.append(row)
        self.listbox.insert("end", self._disp_text(row))
        self.cur_index = len(self.rows)-1
        try:
            self.listbox.selection_clear(0,"end")
            self.listbox.selection_set(self.cur_index)
            self.listbox.see(self.cur_index)
        except: pass
        if hasattr(self, "_zebra_listbox"): self._zebra_listbox()
        self._focus_name1() # カーソルを品目名1へ
        messagebox.showinfo("情報","追加しました")


    # [削除]ボタン
    def on_delete(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("注意","削除対象を選択してください"); return
        del self.rows[idx]; self.listbox.delete(idx)
        self.cur_index=None; self.on_clear()
        if hasattr(self,"_zebra_listbox"): self._zebra_listbox()
        self._focus_name1() # カーソルを品目名1へ
        messagebox.showinfo("情報","削除しました")



    # [更新]ボタン
    def on_update(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("注意","変更対象を左欄から選択してください"); return
        dept = (self.settings or {}).get("department","資材")
        for _,_,csvkey in self.FIELDS:
            if self.schema_is_required(dept, csvkey) and not self.vars.get(csvkey,"").get().strip():
                messagebox.showerror("エラー", f"必須: {csvkey}"); return
        nm1=self.vars.get("品目名1",tk.StringVar(value="")).get()
        nm2=self.vars.get("品目名2",tk.StringVar(value="")).get()
        if not is_ascii_symbol(nm1) or not is_ascii_symbol(nm2):
            messagebox.showerror("エラー","品目名1/2は半角英数記号のみ"); return

        row = self._collect_form_row()
        self.rows[idx] = row
        self.listbox.delete(idx)
        self.listbox.insert(idx, self._disp_text(row))
        try:
            self.listbox.selection_set(idx); self.listbox.see(idx)
        except: pass
        if hasattr(self, "_zebra_listbox"): self._zebra_listbox()
        self._focus_name1() # カーソルを品目名1へ
        messagebox.showinfo("情報","変更しました")


    # [クリア]
    def on_clear(self):
        for _,_,csvkey in self.FIELDS:
            if csvkey in self.vars: self.vars[csvkey].set("")
        self.cur_index=None
        dept = (self.settings or {}).get("department","資材")
        for _,_,k in self.FIELDS:
            m = self.schema_get_mode(dept, k)
            if m in ("fixed","label"):
                v = self.schema_default(dept, k)
                if k=="製造手配区分":
                    self.vars[f"{k}_raw"] = v
                    self.vars[k].set(SEIZO_LABELS.get(v, v))
                elif k in self.vars:
                    self.vars[k].set(v)
        mode_code = self.schema_get_mode(dept, "品目ｺｰﾄﾞ")
        w = self.widgets.get("品目ｺｰﾄﾞ")
        if isinstance(w, tk.Entry):
            w.config(state=("normal" if mode_code == "entry" else "disabled"))
        self._focus_name1()


    # 保存
    def on_save_csv(self):
        if not self.rows:
            messagebox.showwarning("注意","保存対象がありません"); return
        outdir=Path(self.settings.get("output_dir")); outdir.mkdir(parents=True,exist_ok=True)
        path=outdir/"品目新規_エクスポート.csv"
        cols=[csvkey for _,_,csvkey in self.FIELDS]
        with path.open("w",newline="",encoding="utf-8-sig") as f:
            w=csv.DictWriter(f,fieldnames=cols); w.writeheader()
            for r in self.rows: w.writerow({k:r.get(k,"") for k in cols})
        messagebox.showinfo("情報",f"保存しました:\n{path}")

    # 取込
    def on_import_csv(self):
        path=filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
        if not path: return
        cols=[csvkey for _,_,csvkey in self.FIELDS]; tmp=[]
        try:
            with open(path,newline="",encoding="utf-8-sig") as f:
                rdr=csv.DictReader(f)
                for row in rdr: tmp.append({k:row.get(k,"") for k in cols})
        except Exception as e:
            messagebox.showerror("エラー",f"取込失敗: {e}"); return
        self.rows=tmp; self.listbox.delete(0,"end")
        for r in self.rows:
            disp_full=r.get("品目名1") or r.get("品目名2") or r.get("品目ｺｰﾄﾞ") or "(無題)"
            self.listbox.insert("end",trunc31(disp_full))
        self.cur_index=None; self.on_clear(); messagebox.showinfo("情報","取込しました")
        # self._zebra_listbox()
        if hasattr(self, "_zebra_listbox"): self._zebra_listbox()

    # 採番依頼
    def on_request_csv(self):
        if not getattr(self, "rows", None):
            messagebox.showwarning("注意","出力対象がありません"); return
        dept = (self.settings or {}).get("department","資材")
        outdir = Path(self.settings.get("output_dir")); outdir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        path = outdir / f"採番依頼_{ts}.csv"

        cols = getattr(self, "OUT_COLS", [k for _,_,k in self.FIELDS])

        def _default(col:str)->str:
            return self.schema_default(dept, col)

        def _export_value(col:str, val:str)->str:
            m = getattr(self, "dropdown_export_map", {}).get(col, {})
            return (m.get(val, val) if val else "")

        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            for r in self.rows:
                out = {}
                for col in cols:
                    v = r.get(col, "")
                    if not v:
                        v = _default(col)
                    out[col] = _export_value(col, v)
                w.writerow(out)
        messagebox.showinfo("情報", f"採番依頼CSVを出力しました:\n{path}")



if __name__=="__main__":
    App().mainloop()


